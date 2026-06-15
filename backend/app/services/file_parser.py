import io


class ParseResult:
    def __init__(self, total_rows: int, rows: list[dict]):
        self.total_rows = total_rows
        self.rows = rows


class FileParser:
    def parse_bytes(self, content: bytes, filename: str | None = None) -> ParseResult:
        # detect by header bytes for xlsx
        if content[:2] == b"PK":
            # likely xlsx (zip)
            return self._parse_xlsx_bytes(content)
        else:
            return self._parse_csv_bytes(content)

    def _parse_csv_bytes(self, content: bytes) -> ParseResult:
        import csv

        s = io.StringIO(content.decode("utf-8"))
        reader = csv.DictReader(s)
        rows = list(reader)
        return ParseResult(total_rows=len(rows), rows=rows)

    def _parse_xlsx_bytes(self, content: bytes) -> ParseResult:
        from openpyxl import load_workbook

        buf = io.BytesIO(content)
        wb = load_workbook(buf, read_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2):
            rows.append({headers[i]: row[i].value for i in range(len(headers))})
        return ParseResult(total_rows=len(rows), rows=rows)
